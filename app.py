from flask import Flask, render_template, request, send_file, redirect, url_for
import mysql.connector
import xlrd
from copy import copy

from openpyxl.styles import NamedStyle
from datetime import datetime, timedelta
import pytz
import pandas as pd
import calendar
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.fonts import Font
import os
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import PatternFill, GradientFill
import xlwings as xw

app = Flask(__name__)

# Configura la conexión a la base de datos MySQL
conexion_mysql = mysql.connector.connect(
    host='localhost',
    user='root',
    password='',
    database='app_segip'
)

@app.route('/')
def inicio():
    return render_template('inicio.html')

@app.route('/verificar',methods=['GET','POST'])
def verificar():
    if request.method == 'POST':
        fecha_inicio=request.form['fecha_inicio']
        fecha_fin=request.form['fecha_fin']
        cursor_mysql = conexion_mysql.cursor()
        consulta = "select id,fecha,mov,operador,tipo_mov from cabecera WHERE fecha BETWEEN '"+str(fecha_inicio)+"' AND '"+str(fecha_fin)+"' order by fecha ASC , operador asc , tipo_mov DESC"
        cursor_mysql.execute(consulta)   
        resultado = cursor_mysql.fetchall()
         #Convertir los datos a diccionario
        insertObject = []
        columnNames = [column[0] for column in cursor_mysql.description]

        for record in resultado:
            insertObject.append(dict(zip(columnNames, record)))

        conexion_mysql.commit()
        cursor_mysql.close()
        return render_template('verificar.html', data=insertObject,fecha_inicio=fecha_inicio,fecha_fin=fecha_fin)

    else:
        return render_template('verificar.html')
    


@app.route('/eliminar/<string:id>',methods=['GET','POST'])
def eliminar(id):

        cursor_mysql = conexion_mysql.cursor()
        cursor_mysql.execute('DELETE FROM cabecera WHERE id={0}'.format(id))   
        conexion_mysql.commit()

        cursor_mysql = conexion_mysql.cursor()
        cursor_mysql.execute('DELETE FROM detalle WHERE id_cabecera={0}'.format(id))   
        conexion_mysql.commit()
        Flask('Se elimino con exito')
        return redirect(url_for('verificar'))
   



@app.route('/cargar')
def cargar():
    return render_template('cargar.html')

@app.route('/iniciokardex')
def iniciokardex():
    return render_template('iniciokardex.html')



def cerrar_procesos_excel():
    os.system('taskkill /f /im excel.exe')
   

@app.route('/informe',methods=['GET','POST'])
def informe():
    if request.method == 'POST':
        mes=request.form['mes']
        anio=request.form['anio']
        correlativo=request.form['correlativo']
        fila=1

        cerrar_procesos_excel()
       # Conectar a la aplicación Excel
        app = xw.App(visible=False)  # Puedes establecer visible=True si quieres ver Excel

        try:
            wb = xw.Book('informe.xlsx')
            hoja_av4 = wb.sheets['AV-4']
            hoja_av4.range('1:' + str(hoja_av4.api.Rows.Count)).api.Delete()

            dias_del_mes = calendar.monthrange(int(anio), int(mes))[1]
            for dia in range(1, dias_del_mes + 1):
                fecha = datetime(int(anio), int(mes), dia)
                cursor_mysql = conexion_mysql.cursor()
                consulta = "SELECT DISTINCT mov FROM cabecera WHERE fecha = %s order BY operador;"
                #consulta = "SELECT DISTINCT mov FROM cabecera WHERE fecha = %s"
                cursor_mysql.execute(consulta, (fecha,))   
                resultado = cursor_mysql.fetchall()
                conexion_mysql.commit()
                cursor_mysql.close()
                if resultado:
                    rango_copiar = wb.sheets['PLANTILLA'].range('A1:T7')
                    rango_paste = wb.sheets['AV-4'].range('A' + str(fila))  # Especifica la celda de inicio en la hoja de destino
                    rango_copiar.copy(destination=rango_paste)  
                    hoja_av4.range('S' + str(fila + 4)).value = fecha.strftime("%d/%m/%Y")
                    hoja_av4.range('O' + str(fila)).value = 'Correlativo-Form.:   SEGIP/DDSC/MONT/'+ str(correlativo) +'/2024'
                    fila=fila+6
                    inicio=fila+1
                    for row, registro in enumerate(resultado, start=1):
                        cursor_mysql = conexion_mysql.cursor()
                        consulta = "SELECT * FROM cabecera c,detalle d where c.id=d.id_cabecera and c.mov=%s ORDER BY d.descripcion asc, c.tipo_mov desc;"
                        cursor_mysql.execute(consulta,(registro[0],))   
                        reg_detalle = cursor_mysql.fetchall()
                        conexion_mysql.commit()
                        cursor_mysql.close()
                        item_anterior=""
                        fila_utilizado=0
                        fila_sinutilizar=0
                        fila_anulado=0
                        
                        for row2, registro_detalle in enumerate(reg_detalle, start=1):
                                
                            if registro_detalle[10]==item_anterior:
                                if registro_detalle[14]=="":# entrega diaria
                                        fila=fila+1
                                        hoja_av4.range('B' + str(fila)).value = registro_detalle[4]
                                        hoja_av4.range('C' + str(fila)).value = registro_detalle[10]
                                        hoja_av4.range('D' + str(fila)).value = registro_detalle[11]
                                        hoja_av4.range('E' + str(fila)).value = registro_detalle[12]
                                        hoja_av4.range('F' + str(fila)).value = registro_detalle[13]
                                        hoja_av4.range('G' + str(fila)).value = registro_detalle[15]
                                        hoja_av4.range('R' + str(fila)).formula = '=J'+str(fila)+'+M'+str(fila)+'+Q'+str(fila)        
                                if registro_detalle[14]=="MATERIAL UTILIZADO":
                                        if fila_utilizado<=fila:
                                            hoja_av4.range('H' + str(fila_utilizado)).value = registro_detalle[12]
                                            hoja_av4.range('I' + str(fila_utilizado)).value = registro_detalle[13]
                                            hoja_av4.range('J' + str(fila_utilizado)).value = registro_detalle[15]
                                            fila_utilizado=fila_utilizado+1
                                        else:
                                            fila=fila+1
                                            hoja_av4.range('B' + str(fila_utilizado)).value = registro_detalle[4]
                                            hoja_av4.range('C' + str(fila_utilizado)).value = registro_detalle[10]
                                            hoja_av4.range('D' + str(fila_utilizado)).value = registro_detalle[11]
                                            hoja_av4.range('H' + str(fila_utilizado)).value = registro_detalle[12]
                                            hoja_av4.range('I' + str(fila_utilizado)).value = registro_detalle[13]
                                            hoja_av4.range('J' + str(fila_utilizado)).value = registro_detalle[15]
                                            hoja_av4.range('R' + str(fila)).formula = '=J'+str(fila)+'+M'+str(fila)+'+Q'+str(fila)
                                            fila_utilizado=fila_utilizado+1
                                            
                                if registro_detalle[14]=="MATERIAL SIN UTILIZAR":
                                        if fila_sinutilizar<=fila:
                                            hoja_av4.range('O' + str(fila_sinutilizar)).value = registro_detalle[12]
                                            hoja_av4.range('P' + str(fila_sinutilizar)).value = registro_detalle[13]
                                            hoja_av4.range('Q' + str(fila_sinutilizar)).value = registro_detalle[15]
                                            fila_sinutilizar=fila_sinutilizar+1
                                        else:
                                            fila=fila+1
                                            hoja_av4.range('B' + str(fila_sinutilizar)).value = registro_detalle[4]
                                            hoja_av4.range('C' + str(fila_sinutilizar)).value = registro_detalle[10]
                                            hoja_av4.range('D' + str(fila_sinutilizar)).value = registro_detalle[11]
                                            hoja_av4.range('O' + str(fila_sinutilizar)).value = registro_detalle[12]
                                            hoja_av4.range('P' + str(fila_sinutilizar)).value = registro_detalle[13]
                                            hoja_av4.range('Q' + str(fila_sinutilizar)).value = registro_detalle[15]
                                            hoja_av4.range('R' + str(fila)).formula = '=J'+str(fila)+'+M'+str(fila)+'+Q'+str(fila)
                                            fila_sinutilizar=fila_sinutilizar+1
                                            
                                if registro_detalle[14]=="MATERIAL ANULADO":
                                        if fila_anulado<=fila:
                                            hoja_av4.range('K' + str(fila_anulado)).value = registro_detalle[12]
                                            hoja_av4.range('L' + str(fila_anulado)).value = registro_detalle[13]
                                            hoja_av4.range('M' + str(fila_anulado)).value = registro_detalle[15]
                                            fila_anulado=fila_anulado+1
                                        else:
                                            fila=fila+1
                                            hoja_av4.range('B' + str(fila_anulado)).value = registro_detalle[4]
                                            hoja_av4.range('C' + str(fila_anulado)).value = registro_detalle[10]
                                            hoja_av4.range('D' + str(fila_anulado)).value = registro_detalle[11]
                                            hoja_av4.range('K' + str(fila_anulado)).value = registro_detalle[12]
                                            hoja_av4.range('L' + str(fila_anulado)).value = registro_detalle[13]
                                            hoja_av4.range('M' + str(fila_anulado)).value = registro_detalle[15]
                                            hoja_av4.range('R' + str(fila)).formula = '=J'+str(fila)+'+M'+str(fila)+'+Q'+str(fila)
                                            fila_anulado=fila_anulado+1
                                             
                            else:
                                item_anterior=registro_detalle[10]
                                if registro_detalle[14]=="":# entrega diaria
                                        fila=fila+1
                                        fila_utilizado=fila
                                        fila_sinutilizar=fila
                                        fila_anulado=fila
                                        hoja_av4.range('B' + str(fila)).value = registro_detalle[4]
                                        hoja_av4.range('C' + str(fila)).value = registro_detalle[10]
                                        hoja_av4.range('D' + str(fila)).value = registro_detalle[11]
                                        hoja_av4.range('E' + str(fila)).value = registro_detalle[12]
                                        hoja_av4.range('F' + str(fila)).value = registro_detalle[13]
                                        hoja_av4.range('G' + str(fila)).value = registro_detalle[15]
                                        hoja_av4.range('R' + str(fila)).formula = '=J'+str(fila)+'+M'+str(fila)+'+Q'+str(fila)
                    fin=fila
                    fila=fila+1                      
                    rango_copiar = wb.sheets['PLANTILLA'].range('A10:T11')
                    rango_paste = wb.sheets['AV-4'].range('A' + str(fila))  # Especifica la celda de inicio en la hoja de destino
                    rango_copiar.copy(destination=rango_paste)
                    hoja_av4.range('G' + str(fila)).formula = '=SUM(G'+str(inicio)+':G'+str(fin)+')'
                    hoja_av4.range('J' + str(fila)).formula = '=SUM(J'+str(inicio)+':J'+str(fin)+')'
                    hoja_av4.range('M' + str(fila)).formula = '=SUM(M'+str(inicio)+':M'+str(fin)+')'
                    hoja_av4.range('Q' + str(fila)).formula = '=SUM(Q'+str(inicio)+':Q'+str(fin)+')'
                    hoja_av4.range('R' + str(fila)).formula = '=SUM(R'+str(inicio)+':R'+str(fin)+')'
                    hoja_av4.range('S' + str(fila)).formula = '=SUM(S'+str(inicio)+':S'+str(fin)+')'
                    hoja_av4.range('T' + str(fila)).formula = '=SUM(T'+str(inicio)+':T'+str(fin)+')'
                    hoja_av4.range('N' + str(fila)).value = str(inicio)+str(fin)
                    fila=fila+3
                    correlativo=int(correlativo)+1
                    rango = hoja_av4.range('A'+str(inicio)+':T'+str(fin))

                    operador=''
                    color='blanco'
                    for celda in rango:
                        celda.api.Borders(xw.constants.LineStyle.xlContinuous).ColorIndex = 1

    
                
            wb.save('informe.xlsx')
        finally:
            app.quit()
        
        return send_file('informe.xlsx', as_attachment=True)
        
    else:
        return render_template('informe.html')

@app.route('/generar_kardex', methods=['GET','POST'])
def generar_kardex():
    color = 'FFFF00'
    amarillo = PatternFill(start_color=color, end_color=color, fill_type='solid')
    if request.method == 'POST':
        alt_cedula=request.form['alt_cedula']
        alt_lamina=request.form['alt_lamina']
        alt_cedulanueva=request.form['alt_cedulanueva']
        fecha_inicio=request.form['fecha_inicio']
        fecha_fin=request.form['fecha_fin']
        
        if 'archivo' not in request.files:
            return 'No se envió ningún archivo'

        archivo = request.files['archivo']

        if archivo.filename == '':
            return 'No se seleccionó ningún archivo'
        
        wb = load_workbook(archivo)
       
        hoja = wb['KARDEX CI']
        hojalaminas = wb['KARDEX LP']
        hojacedulanueva = wb['KARDEX CI DS4924']


        # comienza para el kardex de ci
        cursor_mysql = conexion_mysql.cursor()
        consulta = "SELECT c.tipo_mov,date_format(c.fecha, '%d/%m/%Y'),c.mov,d.glosa,d.lote,d.desde,d.hasta,d.cantidad FROM cabecera c, detalle d WHERE fecha BETWEEN '"+str(fecha_inicio)+"' AND '"+str(fecha_fin)+"' and c.id=d.id_cabecera AND d.descripcion = 'CEDULAS DE IDENTIDAD' and (d.estado='' or d.estado='MATERIAL SIN UTILIZAR') order by fecha ASC , operador asc , tipo_mov DESC;"
        cursor_mysql.execute(consulta)   
        resultado = cursor_mysql.fetchall()
        conexion_mysql.commit()
        cursor_mysql.close()

        total_filas = hoja.max_row
        #for fila in range(int(alt_cedula) + 1, total_filas + 1):# Eliminar las filas debajo de la fila inicial
        for row_num in range(int(alt_cedula), total_filas + 1):
            hoja.delete_rows(int(alt_cedula))

                    
        for fila, registro in enumerate(resultado, start=1):
            if registro[0] == "ENTREGA DE MATERIAL VALORADO AL OPERADOR":
                valores = [None,str(registro[1]), str(registro[2]),None, str(registro[3]) ,str(registro[4]),None,None,None,int(float(registro[5])),int(float(registro[6])),int(float(registro[7]))]
            if registro[0] == "DEVOLUCIÓN DE MATERIAL VALORADO DEL OPERADOR":
               valores = [None,str(registro[1]), str(registro[2]),None, str(registro[3]) ,str(registro[4]),int(float(registro[5])),int(float(registro[6])),int(float(registro[7]))]     
            
            hoja.append(valores)
            celda_formula = hoja['M' + str(alt_cedula)]
            celda_formula.value = f"=M{str(int(alt_cedula)-1)}+I{str(alt_cedula)}-L{str(alt_cedula)}"  # Ejemplo de fórmula, puedes personalizarla según tus necesidades
            celda_formula.fill = amarillo
            celda = hoja['N' + str(alt_cedula)]
            celda.value="--"
            celda = hoja['O' + str(alt_cedula)]
            celda.value="--"

            celda = hoja['P' + str(alt_cedula)]
            celda.value=f"=M{str(alt_cedula)}"

            celda = hoja['R' + str(alt_cedula)]
            celda.value=f"=I{str(alt_cedula)}*Q{str(alt_cedula)}"

            celda = hoja['S' + str(alt_cedula)]
            celda.value=f"=L{str(alt_cedula)}*Q{str(alt_cedula)}"

            celda = hoja['T' + str(alt_cedula)]
            celda.value=f"=T{str(int(alt_cedula)-1)}+R{str(alt_cedula)}-S{str(alt_cedula)}"
            alt_cedula=int(alt_cedula)+1
        # termina kardex de ci
        

        # comienza para el kardex de lamina
        cursor_mysql = conexion_mysql.cursor()
        consulta = "SELECT c.tipo_mov,date_format(c.fecha, '%d/%m/%Y'),c.mov,d.glosa,d.lote,d.desde,d.hasta,d.cantidad FROM cabecera c, detalle d WHERE fecha BETWEEN '"+str(fecha_inicio)+"' AND '"+str(fecha_fin)+"' and c.id=d.id_cabecera AND d.descripcion = 'LAMINAS PLASTICAS TIPO FUNDA -POUCHE' and (d.estado='' or d.estado='MATERIAL SIN UTILIZAR') order by fecha ASC , operador asc , tipo_mov DESC;"
        cursor_mysql.execute(consulta)   
        resultado = cursor_mysql.fetchall()
        conexion_mysql.commit()
        cursor_mysql.close()

        total_filas = hojalaminas.max_row
        for row_num in range(int(alt_lamina), total_filas + 1):
            hojalaminas.delete_rows(int(alt_lamina))

        for fila, registro in enumerate(resultado, start=1):
            if registro[0] == "ENTREGA DE MATERIAL VALORADO AL OPERADOR":
                valores = [None,str(registro[1]), str(registro[2]),None, str(registro[3]) ,str(registro[4]),None,None,None,int(float(registro[5])),int(float(registro[6])),int(float(registro[7]))]
            if registro[0] == "DEVOLUCIÓN DE MATERIAL VALORADO DEL OPERADOR":
               valores = [None,str(registro[1]), str(registro[2]),None, str(registro[3]) ,str(registro[4]),int(float(registro[5])),int(float(registro[6])),int(float(registro[7]))]     
            
            hojalaminas.append(valores)
            
            celda_formula = hojalaminas['M' + str(alt_lamina)]
            celda_formula.value = f"=M{str(int(alt_lamina)-1)}+I{str(alt_lamina)}-L{str(alt_lamina)}"  # Ejemplo de fórmula, puedes personalizarla según tus necesidades
            celda_formula.fill = amarillo
            celda = hojalaminas['N' + str(alt_lamina)]
            celda.value="--"
            celda = hojalaminas['O' + str(alt_lamina)]
            celda.value="--"

            celda = hojalaminas['P' + str(alt_lamina)]
            celda.value=f"=M{str(alt_lamina)}"

            celda = hojalaminas['R' + str(alt_lamina)]
            celda.value=f"=I{str(alt_lamina)}*Q{str(alt_lamina)}"

            celda = hojalaminas['S' + str(alt_lamina)]
            celda.value=f"=L{str(alt_lamina)}*Q{str(alt_lamina)}"

            celda = hojalaminas['T' + str(alt_lamina)]
            celda.value=f"=T{str(int(alt_lamina)-1)}+R{str(alt_lamina)}-S{str(alt_lamina)}"

            alt_lamina=int(alt_lamina)+1

        # termina kardex de laminas
            
        # comienza para el kardex de cedula DS4924
        cursor_mysql = conexion_mysql.cursor()
        consulta = "SELECT c.tipo_mov,date_format(c.fecha, '%d/%m/%Y'),c.mov,d.glosa,d.lote,d.desde,d.hasta,d.cantidad FROM cabecera c, detalle d WHERE fecha BETWEEN '"+str(fecha_inicio)+"' AND '"+str(fecha_fin)+"' and c.id=d.id_cabecera AND d.descripcion = 'CÉDULA DE IDENTIDAD DS4924' and (d.estado='' or d.estado='MATERIAL SIN UTILIZAR') order by fecha ASC , operador asc , tipo_mov DESC;"
        cursor_mysql.execute(consulta)   
        resultado = cursor_mysql.fetchall()
        conexion_mysql.commit()
        cursor_mysql.close()

        total_filas = hojacedulanueva.max_row
        for row_num in range(int(alt_cedulanueva), total_filas + 1):
            hojacedulanueva.delete_rows(int(alt_cedulanueva))

        for fila, registro in enumerate(resultado, start=1):
            if registro[0] == "ENTREGA DE MATERIAL VALORADO AL OPERADOR":
                valores = [None,str(registro[1]), str(registro[2]),None, str(registro[3]) ,str(registro[4]),None,None,None,int(float(registro[5])),int(float(registro[6])),int(float(registro[7]))]
            if registro[0] == "DEVOLUCIÓN DE MATERIAL VALORADO DEL OPERADOR":
               valores = [None,str(registro[1]), str(registro[2]),None, str(registro[3]) ,str(registro[4]),int(float(registro[5])),int(float(registro[6])),int(float(registro[7]))]     
            
            hojacedulanueva.append(valores)
            celda_formula = hojacedulanueva['M' + str(alt_cedulanueva)]
            celda_formula.value = f"=M{str(int(alt_cedulanueva)-1)}+I{str(alt_cedulanueva)}-L{str(alt_cedulanueva)}"  # Ejemplo de fórmula, puedes personalizarla según tus necesidades
            celda_formula.fill = amarillo

            celda = hojacedulanueva['N' + str(alt_cedulanueva)]
            celda.value="--"
            celda = hojacedulanueva['O' + str(alt_cedulanueva)]
            celda.value="--"

            celda = hojacedulanueva['P' + str(alt_cedulanueva)]
            celda.value=f"=M{str(alt_cedulanueva)}"

            celda = hojacedulanueva['R' + str(alt_cedulanueva)]
            celda.value=f"=I{str(alt_cedulanueva)}*Q{str(alt_cedulanueva)}"

            celda = hojacedulanueva['S' + str(alt_cedulanueva)]
            celda.value=f"=L{str(alt_cedulanueva)}*Q{str(alt_cedulanueva)}"

            celda = hojacedulanueva['T' + str(alt_cedulanueva)]
            celda.value=f"=T{str(int(alt_cedulanueva)-1)}+R{str(alt_cedulanueva)}-S{str(alt_cedulanueva)}"

            alt_cedulanueva=int(alt_cedulanueva)+1
        # termina kardex de cedula DS4924    


        # Guardar los cambios en un nuevo archivo
        archivo_modificado = 'KARDEX.xlsx'
        wb.save(archivo_modificado)

        # Devolver el archivo modificado para su descarga
        return send_file(archivo_modificado, as_attachment=True)
        #return f"altura cedula:{alt_cedula}, altura lamina:{alt_lamina}, altura cedula ds4695: {alt_cedulanueva}"
    


@app.route('/upload', methods=['POST'])
def upload():
    if request.files['archivo']:
        archivo_excel = request.files['archivo']
        workbook = xlrd.open_workbook(file_contents=archivo_excel.read())
        sheet = workbook.sheet_by_index(0)

        cursor_mysql = conexion_mysql.cursor()
        tipo_mov=''
        mov=''
        fecha=''
        operador=''
        for fila in range(1, sheet.nrows):  # Empieza desde la fila 2 para saltar los encabezados
            if fila == 3:   # para el tipo de movimiento
                tipo_mov = sheet.cell_value(fila, 7 )
            if fila == 10:    # para el tipo de codigo de movimiento
                if tipo_mov == "ENTREGA DE MATERIAL VALORADO AL OPERADOR":
                    mov = sheet.cell_value(fila, 11)
                if tipo_mov == "DEVOLUCIÓN DE MATERIAL VALORADO DEL OPERADOR":
                    mov = sheet.cell_value(fila, 12)
            if fila == 13: # para la fecha
                #encabezado[2] = str(sheet.cell_value(fila, 4))
                fecha = xlrd.xldate_as_datetime(sheet.cell_value(fila, 4), workbook.datemode)
            
            #para saber el operador
            if sheet.cell_value(fila, 1)=='ANELY CACERES PECHO' or sheet.cell_value(fila, 1)=='VERONICA MEDRANO ARIAS' or sheet.cell_value(fila, 1)=='IVAR LIMBERT FLORES AYAVIRI' or sheet.cell_value(fila, 1)=='MIGUEL VILLARPANDO MIRANDA' or sheet.cell_value(fila, 1)=='CARMEN DEL PILAR ANTELO PAZ' or sheet.cell_value(fila, 1)=='BOLIVIA MAR PALMERO TILILA' or sheet.cell_value(fila, 1)=='WILSON SOLETO LAVAIN' or sheet.cell_value(fila, 1)=='MIGUEL ANGEL GARCIA ORTEGA' or sheet.cell_value(fila, 1)=='DIEGO ARMANDO YUCRA SILVESTRE' or sheet.cell_value(fila, 1)=='FELIX MARQUINA FERNANDEZ' or sheet.cell_value(fila, 1)=='FRANKLIN POZO HERRERA':
                operador = sheet.cell_value(fila, 1)
        cursor_mysql.execute("SELECT * FROM cabecera WHERE mov=%s AND tipo_mov=%s AND fecha=%s AND operador=%s",(mov,tipo_mov,fecha.strftime('%Y-%m-%d'),operador))
        contador=cursor_mysql.fetchone()

        if contador:
            return render_template('cargar.html',mensaje='Error: El archivo ya existe en la base de datos')
            
        else:
            cursor_mysql.execute("INSERT INTO cabecera (tipo_mov,mov,fecha,operador) VALUES (%s,%s,%s,%s)", (tipo_mov,mov,fecha,operador))
            
            consulta = "SELECT MAX(id) FROM cabecera"
            cursor_mysql.execute(consulta)   
            resultado = cursor_mysql.fetchone()
            ultimo_id = str(resultado[0])

            for fila in range(1, sheet.nrows):
                if sheet.cell_value(fila, 3)=='CEDULAS DE IDENTIDAD' or sheet.cell_value(fila, 3)=='LAMINAS PLASTICAS TIPO FUNDA -POUCHE' or sheet.cell_value(fila, 3)=='CÉDULA DE IDENTIDAD DS4924' :
                    if tipo_mov == "ENTREGA DE MATERIAL VALORADO AL OPERADOR":
                        cursor_mysql.execute("INSERT INTO detalle (id_cabecera,mov,descripcion,lote,desde,hasta,cantidad,glosa) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", (ultimo_id,mov,sheet.cell_value(fila, 3),sheet.cell_value(fila, 8),sheet.cell_value(fila, 12),sheet.cell_value(fila, 13),sheet.cell_value(fila, 15),'ENTREGADO A '+operador))
                    if tipo_mov == "DEVOLUCIÓN DE MATERIAL VALORADO DEL OPERADOR": 
                        cursor_mysql.execute("INSERT INTO detalle (id_cabecera,mov,descripcion,lote,desde,hasta,estado,cantidad,glosa) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)", (ultimo_id,mov,sheet.cell_value(fila, 3),sheet.cell_value(fila, 8),sheet.cell_value(fila, 11),sheet.cell_value(fila, 13),sheet.cell_value(fila, 14),sheet.cell_value(fila, 15),'DEVOLUCION DE '+operador))
            conexion_mysql.commit()
            cursor_mysql.close()
            return render_template('cargar.html',mensaje='El archivo se cargo correctamente')
    else:
        return render_template('cargar.html')



if __name__ == '__main__':
    app.run(debug=True)
